#!/usr/bin/env python3
"""
IFLXI — Importador Excel → PostgreSQL (MVP v1)

NO modifica el Excel del analista.
Lee IFLXI_Carga_Datos_MVP.xlsx (o ruta indicada) y carga en la BD `iflxi`.

Uso (PowerShell, con password ya configurada):
  $env:PGHOST="localhost"
  $env:PGPORT="5432"
  $env:PGUSER="postgres"
  $env:PGPASSWORD="..."
  $env:PGDATABASE="iflxi"

  py importar_excel.py --dry-run
  py importar_excel.py --apply

Opciones:
  --excel RUTA     Excel a leer (por defecto el de esta carpeta)
  --dry-run        Valida y muestra resumen sin escribir
  --apply          Escribe en PostgreSQL
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:
    print("Falta psycopg. Instala: py -m pip install \"psycopg[binary]\" openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "IFLXI_Carga_Datos_MVP.xlsx"
MAP_PATH = ROOT / ".import_map.json"

SHEET_ORDER = [
    "01_Paises",
    "02_Ciudades",
    "03_Competiciones",
    "04_Temporadas",
    "05_Equipos",
    "06_Equipos_Temporada",
    "07_Personas",
    "08_Jugadores",
    "09_Historial",
    "10_Partidos",
    "11_Eventos",
    "12_Fichajes",
    "13_Valor_Mercado",
    "14_Slugs",
]

FORBIDDEN_EVENTS = {"assist", "substitution_in"}


class ImportError_(Exception):
    pass


def v(row: dict, key: str):
    """Valor limpio; None si vacío."""
    if key not in row:
        return None
    x = row[key]
    if x is None:
        return None
    if isinstance(x, str):
        x = x.strip()
        return x if x != "" else None
    return x


def as_bool(x, default=None):
    if x is None:
        return default
    if isinstance(x, bool):
        return x
    s = str(x).strip().lower()
    if s in {"si", "sí", "yes", "true", "1", "y"}:
        return True
    if s in {"no", "false", "0", "n"}:
        return False
    raise ImportError_(f"Valor si/no inválido: {x!r}")


def as_date(x):
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    return date.fromisoformat(s[:10])


def as_int(x):
    if x is None:
        return None
    if isinstance(x, int):
        return x
    return int(float(str(x).strip()))


def as_decimal(x):
    if x is None:
        return None
    if isinstance(x, Decimal):
        return x
    try:
        return Decimal(str(x).strip())
    except InvalidOperation as e:
        raise ImportError_(f"Número inválido: {x!r}") from e


def as_str(x):
    if x is None:
        return None
    return str(x).strip()


def sheet_rows(wb, name: str) -> list[dict]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(as_str(cell.value) if cell.value is not None else None)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        item = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            item[h] = val
            if val is not None and not (isinstance(val, str) and not val.strip()):
                empty = False
        if not empty:
            rows.append(item)
    return rows


def load_map() -> dict:
    if MAP_PATH.exists():
        return json.loads(MAP_PATH.read_text(encoding="utf-8"))
    return {}


def save_map(m: dict) -> None:
    MAP_PATH.write_text(json.dumps(m, indent=2, ensure_ascii=False), encoding="utf-8")


def mid(m: dict, entity: str, code: str | None) -> uuid.UUID | None:
    if not code:
        return None
    bucket = m.setdefault(entity, {})
    if code in bucket:
        return uuid.UUID(bucket[code])
    new_id = uuid.uuid4()
    bucket[code] = str(new_id)
    return new_id


def require(m: dict, entity: str, code: str | None, ctx: str) -> uuid.UUID:
    if not code:
        raise ImportError_(f"{ctx}: falta código {entity}")
    bucket = m.get(entity, {})
    if code not in bucket:
        raise ImportError_(f"{ctx}: no existe {entity}={code}. Cárgalo antes.")
    return uuid.UUID(bucket[code])


def connect():
    conninfo = {
        "host": os.environ.get("PGHOST", "localhost"),
        "port": os.environ.get("PGPORT", "5432"),
        "user": os.environ.get("PGUSER", "postgres"),
        "password": os.environ.get("PGPASSWORD"),
        "dbname": os.environ.get("PGDATABASE", "iflxi"),
    }
    if not conninfo["password"]:
        raise SystemExit(
            "Falta PGPASSWORD. Configura las variables PG* en PowerShell antes de --apply."
        )
    return psycopg.connect(**conninfo, row_factory=dict_row)


def upsert_country(cur, m, row, stats):
    code = as_str(v(row, "codigo_pais"))
    if not code:
        raise ImportError_("01_Paises: codigo_pais obligatorio")
    cid = mid(m, "country", code)
    cur.execute(
        """
        INSERT INTO country (id, iso2, iso3, fifa_code, name_default, continent_code, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (iso2) DO UPDATE SET
          iso3 = EXCLUDED.iso3,
          fifa_code = EXCLUDED.fifa_code,
          name_default = EXCLUDED.name_default,
          continent_code = EXCLUDED.continent_code,
          is_active = EXCLUDED.is_active,
          updated_at = now()
        RETURNING id
        """,
        (
            cid,
            code.upper(),
            as_str(v(row, "codigo_iso3")),
            as_str(v(row, "codigo_fifa")),
            as_str(v(row, "nombre")),
            as_str(v(row, "continente")),
            as_bool(v(row, "activo"), True),
        ),
    )
    real_id = cur.fetchone()["id"]
    m.setdefault("country", {})[code] = str(real_id)
    # también por iso2 normalizado
    m["country"][code.upper()] = str(real_id)
    stats["country"] += 1


def upsert_city(cur, m, row, stats):
    code = as_str(v(row, "codigo_ciudad"))
    if not code:
        raise ImportError_("02_Ciudades: codigo_ciudad obligatorio")
    country_code = as_str(v(row, "codigo_pais"))
    country_id = require(m, "country", country_code, "02_Ciudades")
    city_id = mid(m, "city", code)
    # sin UK natural: delete+insert por id mapeado
    cur.execute(
        """
        INSERT INTO city (id, country_id, name_default, is_active)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (id) DO UPDATE SET
          country_id = EXCLUDED.country_id,
          name_default = EXCLUDED.name_default,
          is_active = EXCLUDED.is_active,
          updated_at = now()
        """,
        (city_id, country_id, as_str(v(row, "nombre")), as_bool(v(row, "activo"), True)),
    )
    stats["city"] += 1


def upsert_competition(cur, m, row, stats):
    code = as_str(v(row, "codigo_competicion"))
    if not code:
        raise ImportError_("03_Competiciones: codigo_competicion obligatorio")
    cid = mid(m, "competition", code)
    country_code = as_str(v(row, "codigo_pais"))
    country_id = require(m, "country", country_code, "03_Competiciones") if country_code else None
    cur.execute(
        """
        INSERT INTO competition (
          id, name_default, short_name, competition_type, scope, country_id,
          governing_body, gender, age_category, is_active
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          name_default = EXCLUDED.name_default,
          short_name = EXCLUDED.short_name,
          competition_type = EXCLUDED.competition_type,
          scope = EXCLUDED.scope,
          country_id = EXCLUDED.country_id,
          governing_body = EXCLUDED.governing_body,
          gender = EXCLUDED.gender,
          age_category = EXCLUDED.age_category,
          is_active = EXCLUDED.is_active,
          updated_at = now()
        """,
        (
            cid,
            as_str(v(row, "nombre")),
            as_str(v(row, "nombre_corto")),
            as_str(v(row, "tipo")),
            as_str(v(row, "ambito")),
            country_id,
            as_str(v(row, "organismo")),
            as_str(v(row, "genero")),
            as_str(v(row, "categoria_edad")) or "senior",
            as_bool(v(row, "activa"), True),
        ),
    )
    stats["competition"] += 1


def upsert_season(cur, m, row, stats):
    code = as_str(v(row, "codigo_temporada"))
    if not code:
        raise ImportError_("04_Temporadas: codigo_temporada obligatorio")
    sid = mid(m, "season", code)
    comp_id = require(m, "competition", as_str(v(row, "codigo_competicion")), "04_Temporadas")
    cur.execute(
        """
        INSERT INTO season (
          id, competition_id, name_default, year_start, year_end,
          start_date, end_date, is_current
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          competition_id = EXCLUDED.competition_id,
          name_default = EXCLUDED.name_default,
          year_start = EXCLUDED.year_start,
          year_end = EXCLUDED.year_end,
          start_date = EXCLUDED.start_date,
          end_date = EXCLUDED.end_date,
          is_current = EXCLUDED.is_current,
          updated_at = now()
        """,
        (
            sid,
            comp_id,
            as_str(v(row, "nombre")),
            as_int(v(row, "anio_inicio")),
            as_int(v(row, "anio_fin")),
            as_date(v(row, "fecha_inicio")),
            as_date(v(row, "fecha_fin")),
            as_bool(v(row, "es_actual"), False),
        ),
    )
    stats["season"] += 1


def upsert_team(cur, m, row, stats):
    code = as_str(v(row, "codigo_equipo"))
    if not code:
        raise ImportError_("05_Equipos: codigo_equipo obligatorio")
    tid = mid(m, "team", code)
    country_id = require(m, "country", as_str(v(row, "codigo_pais")), "05_Equipos")
    city_code = as_str(v(row, "codigo_ciudad"))
    city_id = require(m, "city", city_code, "05_Equipos") if city_code else None
    parent_code = as_str(v(row, "codigo_equipo_padre"))
    parent_id = require(m, "team", parent_code, "05_Equipos") if parent_code else None
    founded = v(row, "anio_fundacion")
    cur.execute(
        """
        INSERT INTO team (
          id, name_default, short_name, code, team_kind, gender, age_category,
          country_id, city_id, parent_team_id, founded_year, is_active
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          name_default = EXCLUDED.name_default,
          short_name = EXCLUDED.short_name,
          code = EXCLUDED.code,
          team_kind = EXCLUDED.team_kind,
          gender = EXCLUDED.gender,
          age_category = EXCLUDED.age_category,
          country_id = EXCLUDED.country_id,
          city_id = EXCLUDED.city_id,
          parent_team_id = EXCLUDED.parent_team_id,
          founded_year = EXCLUDED.founded_year,
          is_active = EXCLUDED.is_active,
          updated_at = now()
        """,
        (
            tid,
            as_str(v(row, "nombre")),
            as_str(v(row, "nombre_corto")),
            as_str(v(row, "codigo_corto")),
            as_str(v(row, "tipo_equipo")),
            as_str(v(row, "genero")),
            as_str(v(row, "categoria_edad")) or "senior",
            country_id,
            city_id,
            parent_id,
            as_int(founded) if founded not in (None, "") else None,
            as_bool(v(row, "activo"), True),
        ),
    )
    stats["team"] += 1


def upsert_team_competition(cur, m, row, stats):
    team_id = require(m, "team", as_str(v(row, "codigo_equipo")), "06_Equipos_Temporada")
    season_id = require(m, "season", as_str(v(row, "codigo_temporada")), "06_Equipos_Temporada")
    code = f"{as_str(v(row, 'codigo_equipo'))}__{as_str(v(row, 'codigo_temporada'))}"
    tcid = mid(m, "team_competition", code)
    cur.execute(
        """
        INSERT INTO team_competition (id, team_id, season_id, status)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (team_id, season_id) DO UPDATE SET
          status = EXCLUDED.status,
          updated_at = now()
        RETURNING id
        """,
        (tcid, team_id, season_id, as_str(v(row, "estado")) or "registered"),
    )
    real = cur.fetchone()["id"]
    m.setdefault("team_competition", {})[code] = str(real)
    stats["team_competition"] += 1


def upsert_person(cur, m, row, stats):
    code = as_str(v(row, "codigo_persona"))
    if not code:
        raise ImportError_("07_Personas: codigo_persona obligatorio")
    pid = mid(m, "person", code)
    bc = as_str(v(row, "codigo_pais_nacimiento"))
    birth_country_id = require(m, "country", bc, "07_Personas") if bc else None
    bcity = as_str(v(row, "codigo_ciudad_nacimiento"))
    birth_city_id = require(m, "city", bcity, "07_Personas") if bcity else None
    cur.execute(
        """
        INSERT INTO person (
          id, full_name, display_name, first_name, last_name,
          birth_date, birth_country_id, birth_city_id, gender
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          full_name = EXCLUDED.full_name,
          display_name = EXCLUDED.display_name,
          first_name = EXCLUDED.first_name,
          last_name = EXCLUDED.last_name,
          birth_date = EXCLUDED.birth_date,
          birth_country_id = EXCLUDED.birth_country_id,
          birth_city_id = EXCLUDED.birth_city_id,
          gender = EXCLUDED.gender,
          updated_at = now()
        """,
        (
            pid,
            as_str(v(row, "nombre_completo")),
            as_str(v(row, "nombre_publico")),
            as_str(v(row, "nombre")),
            as_str(v(row, "apellidos")),
            as_date(v(row, "fecha_nacimiento")),
            birth_country_id,
            birth_city_id,
            as_str(v(row, "genero")),
        ),
    )
    stats["person"] += 1


def upsert_player(cur, m, row, stats):
    code = as_str(v(row, "codigo_jugador"))
    if not code:
        raise ImportError_("08_Jugadores: codigo_jugador obligatorio")
    plid = mid(m, "player", code)
    person_id = require(m, "person", as_str(v(row, "codigo_persona")), "08_Jugadores")
    nat = as_str(v(row, "codigo_pais_nacionalidad"))
    nat_id = require(m, "country", nat, "08_Jugadores") if nat else None
    cur.execute(
        """
        INSERT INTO player (
          id, person_id, nationality_country_id, primary_position, secondary_position,
          foot, height_cm, weight_kg, shirt_name, status
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          person_id = EXCLUDED.person_id,
          nationality_country_id = EXCLUDED.nationality_country_id,
          primary_position = EXCLUDED.primary_position,
          secondary_position = EXCLUDED.secondary_position,
          foot = EXCLUDED.foot,
          height_cm = EXCLUDED.height_cm,
          weight_kg = EXCLUDED.weight_kg,
          shirt_name = EXCLUDED.shirt_name,
          status = EXCLUDED.status,
          updated_at = now()
        """,
        (
            plid,
            person_id,
            nat_id,
            as_str(v(row, "posicion_principal")),
            as_str(v(row, "posicion_secundaria")),
            as_str(v(row, "pie")),
            as_int(v(row, "altura_cm")) if v(row, "altura_cm") is not None else None,
            as_int(v(row, "peso_kg")) if v(row, "peso_kg") is not None else None,
            as_str(v(row, "nombre_camiseta")),
            as_str(v(row, "estado")) or "active",
        ),
    )
    stats["player"] += 1


def upsert_history(cur, m, row, stats):
    code = as_str(v(row, "codigo_historial"))
    if not code:
        raise ImportError_("09_Historial: codigo_historial obligatorio")
    hid = mid(m, "history", code)
    player_id = require(m, "player", as_str(v(row, "codigo_jugador")), "09_Historial")
    team_id = require(m, "team", as_str(v(row, "codigo_equipo")), "09_Historial")
    loan_from = as_str(v(row, "prestado_desde"))
    loan_id = require(m, "team", loan_from, "09_Historial") if loan_from else None
    role = as_str(v(row, "rol")) or "permanent"
    if role == "loan" and not loan_id:
        raise ImportError_(f"09_Historial {code}: rol=loan exige prestado_desde")
    if role != "loan" and loan_id:
        raise ImportError_(f"09_Historial {code}: prestado_desde solo si rol=loan")
    cur.execute(
        """
        INSERT INTO player_team_history (
          id, player_id, team_id, role, start_date, end_date,
          shirt_number, on_loan_from_team_id, notes
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          player_id = EXCLUDED.player_id,
          team_id = EXCLUDED.team_id,
          role = EXCLUDED.role,
          start_date = EXCLUDED.start_date,
          end_date = EXCLUDED.end_date,
          shirt_number = EXCLUDED.shirt_number,
          on_loan_from_team_id = EXCLUDED.on_loan_from_team_id,
          notes = EXCLUDED.notes,
          updated_at = now()
        """,
        (
            hid,
            player_id,
            team_id,
            role,
            as_date(v(row, "fecha_inicio")),
            as_date(v(row, "fecha_fin")),
            as_int(v(row, "dorsal")) if v(row, "dorsal") is not None else None,
            loan_id,
            as_str(v(row, "notas")),
        ),
    )
    stats["history"] += 1


def upsert_match(cur, m, row, stats):
    code = as_str(v(row, "codigo_partido"))
    if not code:
        raise ImportError_("10_Partidos: codigo_partido obligatorio")
    mid_ = mid(m, "match", code)
    season_id = require(m, "season", as_str(v(row, "codigo_temporada")), "10_Partidos")
    home_id = require(m, "team", as_str(v(row, "equipo_local")), "10_Partidos")
    away_id = require(m, "team", as_str(v(row, "equipo_visitante")), "10_Partidos")
    venue = as_str(v(row, "codigo_ciudad_sede"))
    venue_id = require(m, "city", venue, "10_Partidos") if venue else None
    kick = as_str(v(row, "hora_saque_utc"))
    match_date = as_date(v(row, "fecha"))
    kickoff = None
    if kick and match_date:
        # interpreta como UTC naive → timestamptz
        kickoff = datetime.fromisoformat(f"{match_date.isoformat()}T{kick}+00:00")
    cur.execute(
        """
        INSERT INTO match (
          id, season_id, home_team_id, away_team_id, match_date, kickoff_at,
          round_name, status, home_score, away_score, venue_city_id
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          season_id = EXCLUDED.season_id,
          home_team_id = EXCLUDED.home_team_id,
          away_team_id = EXCLUDED.away_team_id,
          match_date = EXCLUDED.match_date,
          kickoff_at = EXCLUDED.kickoff_at,
          round_name = EXCLUDED.round_name,
          status = EXCLUDED.status,
          home_score = EXCLUDED.home_score,
          away_score = EXCLUDED.away_score,
          venue_city_id = EXCLUDED.venue_city_id,
          updated_at = now()
        """,
        (
            mid_,
            season_id,
            home_id,
            away_id,
            match_date,
            kickoff,
            as_str(v(row, "jornada")),
            as_str(v(row, "estado")) or "scheduled",
            as_int(v(row, "goles_local")) if v(row, "goles_local") is not None else None,
            as_int(v(row, "goles_visitante")) if v(row, "goles_visitante") is not None else None,
            venue_id,
        ),
    )
    stats["match"] += 1


def upsert_event(cur, m, row, stats):
    code = as_str(v(row, "codigo_evento"))
    if not code:
        raise ImportError_("11_Eventos: codigo_evento obligatorio")
    eid = mid(m, "event", code)
    etype = as_str(v(row, "tipo_evento"))
    if etype in FORBIDDEN_EVENTS:
        raise ImportError_(f"11_Eventos {code}: tipo prohibido '{etype}' (D-SUB-01 / regla 1)")
    match_id = require(m, "match", as_str(v(row, "codigo_partido")), "11_Eventos")
    player_code = as_str(v(row, "jugador"))
    secondary_code = as_str(v(row, "jugador_secundario"))
    player_id = require(m, "player", player_code, "11_Eventos") if player_code else None
    secondary_id = require(m, "player", secondary_code, "11_Eventos") if secondary_code else None
    team_id = require(m, "team", as_str(v(row, "equipo")), "11_Eventos")
    cur.execute(
        """
        INSERT INTO match_event (
          id, match_id, event_type, player_id, secondary_player_id, team_id,
          minute, extra_minute, period, sort_order
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          match_id = EXCLUDED.match_id,
          event_type = EXCLUDED.event_type,
          player_id = EXCLUDED.player_id,
          secondary_player_id = EXCLUDED.secondary_player_id,
          team_id = EXCLUDED.team_id,
          minute = EXCLUDED.minute,
          extra_minute = EXCLUDED.extra_minute,
          period = EXCLUDED.period,
          sort_order = EXCLUDED.sort_order,
          updated_at = now()
        """,
        (
            eid,
            match_id,
            etype,
            player_id,
            secondary_id,
            team_id,
            as_int(v(row, "minuto")) if v(row, "minuto") is not None else None,
            as_int(v(row, "minuto_extra")) if v(row, "minuto_extra") is not None else None,
            as_str(v(row, "periodo")),
            as_int(v(row, "orden")) if v(row, "orden") is not None else None,
        ),
    )
    stats["event"] += 1


def upsert_transfer(cur, m, row, stats):
    code = as_str(v(row, "codigo_fichaje"))
    if not code:
        raise ImportError_("12_Fichajes: codigo_fichaje obligatorio")
    tid = mid(m, "transfer", code)
    player_id = require(m, "player", as_str(v(row, "codigo_jugador")), "12_Fichajes")
    fr = as_str(v(row, "equipo_origen"))
    to = as_str(v(row, "equipo_destino"))
    from_id = require(m, "team", fr, "12_Fichajes") if fr else None
    to_id = require(m, "team", to, "12_Fichajes") if to else None
    hist = as_str(v(row, "codigo_historial_destino"))
    hist_id = require(m, "history", hist, "12_Fichajes") if hist else None
    fee = v(row, "importe")
    fee_amount = as_decimal(fee) if fee not in (None, "") else None
    ttype = as_str(v(row, "tipo"))
    if ttype in {"free", "end_of_contract"} and fee_amount is not None:
        raise ImportError_(f"12_Fichajes {code}: free/end_of_contract ⇒ importe vacío")
    cur.execute(
        """
        INSERT INTO transfer (
          id, player_id, from_team_id, to_team_id, transfer_type,
          announced_date, effective_date, fee_amount, fee_currency,
          fee_is_estimated, related_history_id
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          player_id = EXCLUDED.player_id,
          from_team_id = EXCLUDED.from_team_id,
          to_team_id = EXCLUDED.to_team_id,
          transfer_type = EXCLUDED.transfer_type,
          announced_date = EXCLUDED.announced_date,
          effective_date = EXCLUDED.effective_date,
          fee_amount = EXCLUDED.fee_amount,
          fee_currency = EXCLUDED.fee_currency,
          fee_is_estimated = EXCLUDED.fee_is_estimated,
          related_history_id = EXCLUDED.related_history_id,
          updated_at = now()
        """,
        (
            tid,
            player_id,
            from_id,
            to_id,
            ttype,
            as_date(v(row, "fecha_anuncio")),
            as_date(v(row, "fecha_efectiva")),
            fee_amount,
            as_str(v(row, "moneda")) if fee_amount is not None else None,
            as_bool(v(row, "importe_estimado"), False),
            hist_id,
        ),
    )
    stats["transfer"] += 1


def upsert_market(cur, m, row, stats):
    code = as_str(v(row, "codigo_valor"))
    if not code:
        raise ImportError_("13_Valor_Mercado: codigo_valor obligatorio")
    mvid = mid(m, "market", code)
    player_id = require(m, "player", as_str(v(row, "codigo_jugador")), "13_Valor_Mercado")
    cur.execute(
        """
        INSERT INTO market_value_history (
          id, player_id, value_amount, currency, recorded_on, source
        ) VALUES (%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          player_id = EXCLUDED.player_id,
          value_amount = EXCLUDED.value_amount,
          currency = EXCLUDED.currency,
          recorded_on = EXCLUDED.recorded_on,
          source = EXCLUDED.source
        """,
        (
            mvid,
            player_id,
            as_decimal(v(row, "importe")),
            as_str(v(row, "moneda")),
            as_date(v(row, "fecha_valoracion")),
            as_str(v(row, "fuente")) or "manual",
        ),
    )
    stats["market"] += 1


def upsert_slug(cur, m, row, stats):
    code = as_str(v(row, "codigo_slug"))
    if not code:
        raise ImportError_("14_Slugs: codigo_slug obligatorio")
    sid = mid(m, "slug", code)
    etype = as_str(v(row, "tipo_entidad"))
    ecode = as_str(v(row, "codigo_entidad"))
    entity_map = {
        "player": "player",
        "team": "team",
        "competition": "competition",
        "match": "match",
        "transfer": "transfer",
        "season": "season",
    }
    if etype not in entity_map:
        raise ImportError_(f"14_Slugs {code}: tipo_entidad inválido {etype}")
    entity_id = require(m, entity_map[etype], ecode, "14_Slugs")
    cur.execute(
        """
        INSERT INTO slug (
          id, entity_type, entity_id, locale, slug, is_primary, is_active
        ) VALUES (%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id) DO UPDATE SET
          entity_type = EXCLUDED.entity_type,
          entity_id = EXCLUDED.entity_id,
          locale = EXCLUDED.locale,
          slug = EXCLUDED.slug,
          is_primary = EXCLUDED.is_primary,
          is_active = EXCLUDED.is_active,
          updated_at = now()
        """,
        (
            sid,
            etype,
            entity_id,
            as_str(v(row, "idioma")),
            as_str(v(row, "slug")),
            as_bool(v(row, "es_principal"), True),
            as_bool(v(row, "activo"), True),
        ),
    )
    stats["slug"] += 1


def refresh_caches(cur):
    """Actualiza caches de player desde HISTORY (club abierto) y MARKET_VALUE."""
    cur.execute(
        """
        UPDATE player p
        SET current_team_id = s.team_id,
            updated_at = now()
        FROM (
          SELECT DISTINCT ON (h.player_id)
            h.player_id, h.team_id
          FROM player_team_history h
          JOIN team t ON t.id = h.team_id
          WHERE h.end_date IS NULL
            AND t.team_kind = 'club'
          ORDER BY h.player_id, h.start_date DESC
        ) s
        WHERE p.id = s.player_id
        """
    )
    cur.execute(
        """
        UPDATE player p
        SET current_team_id = NULL,
            updated_at = now()
        WHERE NOT EXISTS (
          SELECT 1
          FROM player_team_history h
          JOIN team t ON t.id = h.team_id
          WHERE h.player_id = p.id
            AND h.end_date IS NULL
            AND t.team_kind = 'club'
        )
        """
    )
    cur.execute(
        """
        UPDATE player p
        SET current_market_value = m.value_amount,
            current_market_value_currency = m.currency,
            updated_at = now()
        FROM (
          SELECT DISTINCT ON (player_id)
            player_id, value_amount, currency
          FROM market_value_history
          ORDER BY player_id, recorded_on DESC, created_at DESC
        ) m
        WHERE p.id = m.player_id
        """
    )


HANDLERS = {
    "01_Paises": upsert_country,
    "02_Ciudades": upsert_city,
    "03_Competiciones": upsert_competition,
    "04_Temporadas": upsert_season,
    "05_Equipos": upsert_team,
    "06_Equipos_Temporada": upsert_team_competition,
    "07_Personas": upsert_person,
    "08_Jugadores": upsert_player,
    "09_Historial": upsert_history,
    "10_Partidos": upsert_match,
    "11_Eventos": upsert_event,
    "12_Fichajes": upsert_transfer,
    "13_Valor_Mercado": upsert_market,
    "14_Slugs": upsert_slug,
}


def main():
    ap = argparse.ArgumentParser(description="IFLXI importador Excel → PostgreSQL")
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    if args.dry_run == args.apply:
        print("Indica exactamente una opción: --dry-run  o  --apply")
        sys.exit(2)

    if not args.excel.exists():
        print(f"No encuentro el Excel: {args.excel}")
        sys.exit(1)

    wb = load_workbook(args.excel, data_only=True)
    m = load_map()
    stats = {k: 0 for k in [
        "country", "city", "competition", "season", "team", "team_competition",
        "person", "player", "history", "match", "event", "transfer", "market", "slug",
    ]}
    errors: list[str] = []

    print(f"Excel: {args.excel}")
    print(f"Modo: {'DRY-RUN' if args.dry_run else 'APPLY'}")
    print(f"Hojas: {', '.join(s for s in SHEET_ORDER if s in wb.sheetnames)}")

    if args.dry_run:
        defined_countries: set[str] = set()
        used_countries: list[tuple[str, str, str]] = []
        for sheet in SHEET_ORDER:
            rows = sheet_rows(wb, sheet)
            print(f"  {sheet}: {len(rows)} filas con datos")
            for i, row in enumerate(rows, start=2):
                try:
                    if sheet == "01_Paises":
                        c = as_str(v(row, "codigo_pais"))
                        if not c:
                            raise ImportError_("codigo_pais vacío")
                        defined_countries.add(c.upper())
                    if sheet == "11_Eventos":
                        et = as_str(v(row, "tipo_evento"))
                        if et in FORBIDDEN_EVENTS:
                            raise ImportError_(f"tipo prohibido {et}")
                    if sheet == "09_Historial":
                        role = as_str(v(row, "rol")) or "permanent"
                        loan = as_str(v(row, "prestado_desde"))
                        if role == "loan" and not loan:
                            raise ImportError_("loan sin prestado_desde")
                        if role != "loan" and loan:
                            raise ImportError_("prestado_desde con rol distinto de loan")
                    # países referenciados
                    for col in (
                        "codigo_pais",
                        "codigo_pais_nacimiento",
                        "codigo_pais_nacionalidad",
                    ):
                        c = as_str(v(row, col))
                        if c and sheet != "01_Paises":
                            used_countries.append((sheet, col, c.upper()))
                except Exception as e:
                    errors.append(f"{sheet} fila~{i}: {e}")
        for sheet, col, c in used_countries:
            if c not in defined_countries:
                errors.append(
                    f"{sheet}: usa país {c} en {col}, pero no está en 01_Paises"
                )
        if errors:
            print("\nErrores de validación:")
            for e in errors:
                print(" -", e)
            sys.exit(1)
        print("\nDRY-RUN OK. Sin escritura en BD.")
        print("Cuando quieras cargar: py importar_excel.py --apply")
        return

    # APPLY
    with connect() as conn:
        with conn.transaction():
            with conn.cursor() as cur:
                for sheet in SHEET_ORDER:
                    rows = sheet_rows(wb, sheet)
                    handler = HANDLERS[sheet]
                    for i, row in enumerate(rows, start=2):
                        try:
                            handler(cur, m, row, stats)
                        except Exception as e:
                            raise ImportError_(f"{sheet} fila~{i}: {e}") from e
                refresh_caches(cur)
        save_map(m)

    print("\nIMPORTACIÓN OK")
    for k, n in stats.items():
        if n:
            print(f"  {k}: {n}")
    print(f"Mapa códigos→UUID: {MAP_PATH}")
    print("Siguiente: ejecuta sql/IFLXI-validaciones-carga.sql")


if __name__ == "__main__":
    try:
        main()
    except ImportError_ as e:
        print("ERROR:", e)
        sys.exit(1)
